import requests
import re
import os

import urllib
from bs4 import BeautifulSoup

from lxml import etree

import openpyxl



import traceback
import datetime
import getpass

# date
date = datetime.datetime.now().strftime("%x")
URL_PORT_SUM = "https://alger.olaccess2.com/rest/1.0/PortfolioSummary/"
URL_LOGON = "https://alger.olaccess2.com/logon/rest/1.0/LoginWizard/password"
file_path = "Alger Virtual Portfolio Manager\Finance.xlsx"

# Author Daniel Martin


import sys
import threading
import itertools
import time
from openpyxl import Workbook, cell
import math

class Spinner:

    def __init__(self, message="", delay=0.1):
        self.spinner = itertools.cycle(['-', '/', '|', '\\'])
        self.delay = delay
        self.busy = False
        self.spinner_visible = False
        sys.stdout.write(message)

    def write_next(self):
        with self._screen_lock:
            if not self.spinner_visible:
                sys.stdout.write(next(self.spinner))
                self.spinner_visible = True
                sys.stdout.flush()

    def remove_spinner(self, cleanup=False):
        with self._screen_lock:
            if self.spinner_visible:
                sys.stdout.write('\b')
                self.spinner_visible = False
                if cleanup:
                    sys.stdout.write(' ')       # overwrite spinner with blank
                    sys.stdout.write('\r')      # move to next line
                sys.stdout.flush()

    def spinner_task(self):
        while self.busy:
            self.write_next()
            time.sleep(self.delay)
            self.remove_spinner()

    def __enter__(self):
        if sys.stdout.isatty():
            self._screen_lock = threading.Lock()
            self.busy = True
            self.thread = threading.Thread(target=self.spinner_task)
            self.thread.start()
            return self

    def __exit__(self, exception, value, tb):
        if sys.stdout.isatty():
            self.busy = False
            self.remove_spinner(cleanup=True)
        else:
            sys.stdout.write('\r')


def listToString(s):

    # initialize an empty string
    str1 = ""

    # traverse in the string
    for element in s:
        str1 += element

    # return string
    return str1


def find_between(s, first, last):
    try:

        start = s.index(first) + len(first)
        end = s.index(last, start)
        return s[start:end]

    except ValueError:
        return ""


def find_between_r(s, first, last):
    try:
        start = s.rindex(first) + len(first)
        end = s.rindex(last, start)
        return s[start:end]
    except ValueError:
        return ""


def delete_previous_line():
    '''
    Deletes previous print statement
    '''
    os.system("echo \033[2A")
    os.system("echo \033[2K")


def website_login_and_retrieve_info():
    '''
    "Accesses Alger Fund's logon page and retrieves portfolio information"
    @return: list[]
    '''
    
    

    with requests.Session() as s:
        try:
            while True:
                loginName = str(input("Enter User ID: "))
                psw = getpass.getpass()
                AuthCode = str(input("Input One-Time PIN Code: "))

                values = {"loginName": loginName, "password": psw, 'otp': AuthCode}

                with Spinner("Validating login standby...") as spinner:
                        # POST login info
                        r = s.post(URL_LOGON, json=values, allow_redirects=True)
                        # GET portfolio summary page
                        p = s.get(URL_PORT_SUM)
                        # Authorization required going forward
                        # if user not authenticated
                        
                        if p.status_code & r.status_code != 200:
                            print("ERROR-> Login Failed. Please check credentials and try again")
                            print(
                                '''If you believe this to be an error, please contact Alger Customer Service at
                                1-800-992-3863
                                Mon - Fri between 7am - 7pm CST
                            ''')
                            
                            r.close()
                            p.close()
                            spinner.__exit__(Exception,0,0)
                            continue
                        elif p.status_code & r.status_code == 200:
                            break
            with Spinner("Collecting Information"):
                delete_previous_line()
                # get response
                soup = BeautifulSoup(p.text, features='lxml')
                response = soup.find('p')
                response = listToString(response.contents)

                # extrapolate certian values from port_sum pages to build GET REQUEST URL
                masterRegId = find_between(response, '"masterregid":', ',')
                acctType = find_between(response, '"acctType":', ',')
                acctNbr = find_between(response, '"acctNbr":', ',')
                cdOwner = find_between(response, '"cdOwner":', ',')

                # Build GET REQUEST Uniform Resource Loader with above-obtained values
                URL_ACCOUNT_LISTING = f"https://alger.olaccess2.com/rest/1.0/AccountListing/getAccountListing?masterRegId={masterRegId}&acctType={acctType}&cdOwner={cdOwner}&acctNbr={acctNbr}"
                # GET request for account listing information
                m = s.get(URL_ACCOUNT_LISTING)

                print("Account accessed sucessfully. Extracting Data\r")
                delete_previous_line()

                # get response
                soup = BeautifulSoup(m.text, features="lxml")
                response = soup.find('p')
                response = listToString(response.contents)

                # get the following values from the GET request URL
                value = find_between(response, '"acctValue":', ',')
                fundName = find_between(response, '"nameFund":', ',')
                shares = find_between(response, '"shareBalance":', ',')
                sharePrice = find_between(response, '"amtPrceStat":', ',')


            r.close()
            p.close()
            m.close()
            return [fundName, float(value), float(shares), float(sharePrice)]

        except Exception as e:
            sys.stdout.flush()
            r.close()
            p.close()
            s.close()
            print("Unexpected error:", sys.exc_info()[0])
            raise
from openpyxl.utils import get_column_letter
def create_excelsheet():
    with Spinner("Creating excelsheet...") as spinner:
        try:
            wb = Workbook()
            sheet = wb.active
            sheet['A1'] = "Date"
            sheet['B1'] = "Shares"
            sheet['C1'] = "Share Price"
            sheet['D1'] = "Value"
            sheet['E1'] = "Original % Growth"
            sheet['F1'] = "% Growth since Last Entry"
        
            
            for cell in sheet[1]:
                col = cell.column_letter
                length = len(str(cell.value))
                sheet.column_dimensions['{}'.format(col)].width = length + 4
                
            
            save_excelsheet(workbook=wb)    
                
        except TypeError as e:
            print("An error occured...")
            save_excelsheet(workbook=wb)
            traceback.print_exception(e)
                
def open_excelsheet(filename=file_path, readOnly=False):
    with Spinner():
        try:
            workbook = openpyxl.load_workbook(filename, data_only=readOnly)
        except PermissionError:
            open(file_path).close()
            workbook = openpyxl.load_workbook(filename, data_only=readOnly)
    print("Success: Accessed excelsheet\r")
    delete_previous_line()
    return workbook


def save_excelsheet(filename=file_path, workbook=Workbook()):
    with Spinner():
        workbook.save(filename)
        workbook.close()
    print("success: saved and closed excelsheet\r")
    delete_previous_line()



def update_excelsheet(fundName=None, value=None, tot_shares=None, price_share=None):
    '''
    Amends excelsheet containing portfolio records
    '''
    try: 
        os.rename(file_path, 'tempfile.xls')
        os.rename('tempfile.xls', file_path)
    except OSError:
        while True:
            print("Detected another process running the current excelsheet. Please close file to continue: " + file_path)
            time.sleep(2)
            delete_previous_line()
            os.system("echo \033[2A")
            try:
                os.rename(file_path, 'tempfile.xls')
                os.rename('tempfile.xls', file_path)
                break
            except OSError:
                continue

    workbook = open_excelsheet()
    sheet = workbook.active
    

    last_entry_date = None

    #Example of bad design from openpyxl 0/5 seriously It's actually the worst
    #Basically if user deletes a row of data, we need to detect it
    for cell in sheet['D']:
        if not cell.value:
            sheet.delete_rows(cell.row,1)

    sheet.append([date,tot_shares,price_share,value])
    current_row = sheet.max_row
    print(current_row)
    if not int(current_row):
        print("There seems to be a problem")
        os._exit(0)
    if current_row == 2:
        # set %Orig Growth & set %last entry growth
        sheet['E2'].style = 'Percent'
        sheet['F2'].style = 'Percent'
        last_growth = orig_growth = sheet['F2'].value = sheet['E2'].value = 0
            
    else:
        next_row=current_row+1
        prev_row=current_row-1

        # get original entry portfolio value
        original = sheet['D2'].value

        # calculate and set % growth since original
        sheet[f'E{current_row}'].style = 'Percent'
        orig_growth = sheet[f'E{current_row}'] = round(float((value - original) / original),3)
        orig_growth= abs(round(orig_growth*100,3))
        
        # set and calculate %last entry growth
        last_val = sheet[f'D{prev_row}'].value
        sheet[f'F{current_row}'].style = 'Percent'

        last_growth = sheet[f'F{current_row}'] = round(float((value - last_val)/last_val),3)
        last_growth = abs(round(last_growth*100,3))
        if sheet[f'A{prev_row}']:
            last_entry_date = sheet[f'A{prev_row}'].value
        else:
            last_entry_date = date


    save_excelsheet(workbook=workbook)
    # Create a dict to store records
    financeRecord = {
        "Date ": date,
        "Shares ": tot_shares,
        "Price of Share ": price_share,
        "Value ": value,
        "Growth Since Beginning": float(orig_growth),
        "Growth Since Last Entry": float(last_growth),
        "Fund Name": fundName}
    
    if last_growth > 0:
        phrase = "increase"
    elif last_growth < 0:
        phrase = "decrease"
    if orig_growth > 0:
        phrase1 = "increased"
    elif orig_growth < 0:
        phrase1 = "decreased"
    os.system(('cls'))
    if last_entry_date:
        print(
            f'''
            Welcome to Alger Virtual Portfolio Manager. 
            Today is {financeRecord['Date ']}. 
            Your last entry date was {last_entry_date}.
            Here are your records for your {financeRecord['Fund Name']}
            Your total number of shares are {financeRecord['Shares ']}, which are worth ${financeRecord['Price of Share ']} indvidually.
            Your total portfolio value is ${financeRecord['Value ']}.
            {"Your portfolio value has not grown since your last entry" if financeRecord['Growth Since Last Entry']==0 else f"This is a {financeRecord['Growth Since Last Entry']}% {phrase} since your last entry."}
            Since your first entry, your portfolio value has {"not grown." if financeRecord['Growth Since Beginning']==0 else f"{phrase1} by {financeRecord['Growth Since Beginning']}%"}
            You may access these records and more at any time in your files
            Thank you. 
            '''
        )
    elif not last_entry_date:
          print(
            f'''
            Welcome to Alger Virtual Portfolio Manager. 
            Today is {financeRecord["Date "]}. 
            Our records indicate this is your first entry. Welcome!
            Here are your records for your {financeRecord["Fund Name"]}
            Your total number of shares are {financeRecord['Shares ']}, which are worth ${financeRecord['Price of Share ']} indvidually.
            Your total portfolio value is ${financeRecord['Value ']}.
            You may access these records and more at any time in your files
            Thank you.
            '''
          )
    sys.stdout.write("Goodbye :)"+ "v: 1.5.9")
    sys.stdout.flush()

info = website_login_and_retrieve_info()

try:
    info[1]=round(info[1],2)
except Exception as e:
    print("Error occured; Could not extract data properly")
    traceback.print(e)
print("Data extracted successfully...\r")
delete_previous_line()

#if file exists
if os.path.exists(file_path):
    update_excelsheet(info[0],info[1],info[2],info[3])
#otherwise create file
else:
    create_excelsheet()
    update_excelsheet(info[0],info[1],info[2],info[3])
